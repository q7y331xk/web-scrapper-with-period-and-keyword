from openpyxl import Workbook
from config import EXCEL_SAVE_PATH

def write_excel(sellings):
    col_length = len(sellings)
    row_length = len(sellings[0])
    wb = Workbook()
    ws = wb.active
    ws.title = "sellings"
    ws.cell(row = 1, column = 1, value = "id")
    ws.cell(row = 1, column = 2, value = "title")
    ws.cell(row = 1, column = 3, value = "status")
    ws.cell(row = 1, column = 4, value = "views")
    ws.cell(row = 1, column = 5, value = "likes")
    ws.cell(row = 1, column = 6, value = "comments")
    ws.cell(row = 1, column = 7, value = "name")
    ws.cell(row = 1, column = 8, value = "date")
    i = 0
    while (i < col_length):
        row = sellings[i]
        j = 0
        while (j < row_length):
            ws.cell(row = i + 1 + 1, column = j + 1, value = row[j])
            j = j + 1
        i = i + 1
    wb.save(EXCEL_SAVE_PATH)
    print(f'write excel in "{EXCEL_SAVE_PATH}"')